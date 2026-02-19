"""
Backend API for Multi-Questionnaire System
Supports multiple questionnaires loaded from JSON files
"""
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime
import json
import os
import uuid
import glob
import io
from analysis_engine import AnalysisEngine
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Paths
BACKEND_DIR = os.path.dirname(__file__)
FRONTEND_DIR = os.path.join(os.path.dirname(BACKEND_DIR), "frontend")
QUESTIONNAIRES_DIR = os.path.join(BACKEND_DIR, "questionnaires")
DATA_DIR = os.path.join(BACKEND_DIR, "data")
RESULTS_PDF_DIR = os.path.join(DATA_DIR, "resultados_pdf")

app = FastAPI(
    title="Sistema de Cuestionarios",
    description="API for managing multiple questionnaires and responses",
    version="2.0.0",
    root_path="/cuestionarios"
)

# Ensure directories exist
os.makedirs(RESULTS_PDF_DIR, exist_ok=True)

from fastapi import UploadFile, File
import shutil

# Endpoint eliminado: save-pdf (se reemplazó por generate-pdf-server con Playwright)


class PDFGenerationRequest(BaseModel):
    respondent_cedula: str
    submitted_at: str
    responses: List[Dict[str, Any]]


def build_report_html(data: PDFGenerationRequest, img_colombia: str, img_javeriana: str) -> str:
    """Build the full HTML for the stress report, identical to the frontend version."""
    questions = [
        (1, "Dolores en el cuello y espalda o tensión muscular."),
        (2, "Problemas gastrointestinales, úlcera péptica, acidez, problemas digestivos o del colon."),
        (3, "Problemas respiratorios."),
        (4, "Dolor de cabeza."),
        (5, "Trastornos del sueño como somnolencia durante el día o desvelo en la noche."),
        (6, "Palpitaciones en el pecho o problemas cardíacos."),
        (7, "Cambios fuertes del apetito."),
        (8, "Problemas relacionados con la función de los órganos genitales (impotencia, frigidez)."),
        (9, "Dificultad en las relaciones familiares."),
        (10, "Dificultad para permanecer quieto o dificultad para iniciar actividades."),
        (11, "Dificultad en las relaciones con otras personas."),
        (12, "Sensación de aislamiento y desinterés."),
        (13, "Sentimiento de sobrecarga de trabajo."),
        (14, "Dificultad para concentrarse, olvidos frecuentes."),
        (15, "Aumento en el número de accidentes de trabajo."),
        (16, "Sentimiento de frustración, de no haber hecho lo que se quería en la vida."),
        (17, "Cansancio, tedio o desgano."),
        (18, "Disminución del rendimiento en el trabajo o poca creatividad."),
        (19, "Deseo de no asistir al trabajo."),
        (20, "Bajo compromiso o poco interés con lo que se hace."),
        (21, "Dificultad para tomar decisiones."),
        (22, "Deseo de cambiar de empleo."),
        (23, "Sentimiento de soledad y miedo."),
        (24, "Sentimiento de irritabilidad, actitudes y pensamientos negativos."),
        (25, "Sentimiento de angustia, preocupación o tristeza."),
        (26, "Consumo de drogas para aliviar la tensión o los nervios."),
        (27, 'Sentimientos de que "no vale nada" o "no sirve para nada".'),
        (28, "Consumo de bebidas alcohólicas o café o cigarrillo."),
        (29, "Sentimiento de que está perdiendo la razón."),
        (30, "Comportamientos rígidos, obstinación o terquedad."),
        (31, "Sensación de no poder manejar los problemas de la vida."),
    ]

    responses_map = {r["question_id"]: r["response_value"] for r in data.responses}

    from datetime import datetime as dt
    submission_date = dt.fromisoformat(data.submitted_at.replace("Z", "+00:00"))
    day = str(submission_date.day).zfill(2)
    month = str(submission_date.month).zfill(2)
    year = submission_date.year

    table_rows = ""
    for qid, qtext in questions:
        val = responses_map.get(qid, 0)
        table_rows += f"""
        <tr>
            <td class="question-cell"><span class="question-number">{qid}.</span> {qtext}</td>
            <td class="response-cell">{"X" if val == 1 else ""}</td>
            <td class="response-cell">{"X" if val == 2 else ""}</td>
            <td class="response-cell">{"X" if val == 3 else ""}</td>
            <td class="response-cell">{"X" if val == 4 else ""}</td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Reporte - CC: {data.respondent_cedula}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}

        @page {{
            size: letter;
            margin: 1.5cm 2cm;
        }}

        body {{
            font-family: Arial, sans-serif;
            font-size: 11pt;
            line-height: 1.3;
            color: #000;
            background: white;
        }}

        /* === CARÁTULA === */
        .cover-page {{
            page-break-after: always;
        }}

        .cover-header-table {{
            width: 100%;
            margin-bottom: 3cm;
            margin-top: 1cm;
        }}

        .cover-header-table td {{
            vertical-align: middle;
            padding: 5px;
        }}

        .cover-data-label {{
            font-size: 10pt;
            text-align: right;
            width: 200px;
        }}

        .box {{
            border: 1pt solid #000;
            height: 28px;
            text-align: center;
            vertical-align: middle;
            font-size: 10pt;
            padding: 2px 4px;
        }}

        .day-box, .month-box {{ width: 30px; }}
        .year-box {{ width: 60px; }}
        .id-box {{ width: 120px; }}

        .box-sublabel {{
            font-size: 7pt;
            color: #666;
            text-align: center;
        }}

        .cover-main {{
            text-align: center;
            margin-top: 5cm;
            margin-bottom: 5cm;
        }}

        .green-title {{
            color: #2e7d32;
            font-size: 15pt;
            font-weight: bold;
            text-transform: uppercase;
            line-height: 1.5;
        }}

        .cover-footer-table {{
            width: 100%;
            margin-top: 2cm;
        }}

        .cover-footer-table td {{
            vertical-align: bottom;
            text-align: center;
        }}

        .footer-logo-left img {{ height: 110px; }}
        .footer-logo-left .org-name {{ font-size: 9pt; font-weight: bold; display: block; }}
        .footer-logo-left .org-sub {{ font-size: 9pt; display: block; }}
        .footer-logo-right img {{ height: 106px; }}

        /* === PÁGINA DE CONTENIDO === */
        .header-table {{
            width: 100%;
            margin-bottom: 5px;
        }}

        .header-table td {{
            vertical-align: middle;
        }}

        .logo-left-cell {{ width: 50%; text-align: left; }}
        .logo-right-cell {{ width: 50%; text-align: right; }}
        .logo-left-cell img, .logo-right-cell img {{ height: 65px; width: auto; }}

        .logo-text {{ font-size: 9pt; line-height: 1.2; vertical-align: middle; padding-left: 8px; }}
        .logo-text strong {{ display: block; font-size: 10pt; }}

        .title {{
            text-align: center;
            font-size: 10.5pt;
            font-weight: bold;
            margin: 5px 0;
            text-transform: uppercase;
        }}

        .instructions {{
            margin-bottom: 5px;
            font-size: 9pt;
            text-align: justify;
        }}

        .questionnaire-table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
        }}

        .questionnaire-table th {{
            border: 1pt solid #000;
            padding: 4px 5px;
            text-align: center;
            font-size: 9pt;
            font-weight: bold;
            background-color: #d8d8d8;
        }}

        .questionnaire-table td {{
            border: 1pt solid #000;
            padding: 2px 8px;
            font-size: 9pt;
            vertical-align: middle;
        }}

        .question-cell {{ text-align: left; width: 55%; }}
        .response-cell {{ text-align: center; width: 11.25%; font-size: 11pt; font-weight: bold; }}
        .question-number {{ font-weight: normal; margin-right: 5px; }}
    </style>
</head>
<body>

    <!-- CARÁTULA -->
    <div class="cover-page">

        <!-- Fecha e ID alineados a la derecha -->
        <table class="cover-header-table">
            <tr>
                <td class="cover-data-label">Fecha de aplicación:</td>
                <td>
                    <table><tr>
                        <td>
                            <table><tr>
                                <td class="box day-box">{day}</td>
                                <td class="box month-box">{month}</td>
                                <td class="box year-box">{year}</td>
                            </tr><tr>
                                <td class="box-sublabel">dd</td>
                                <td class="box-sublabel">mm</td>
                                <td class="box-sublabel">aaaa</td>
                            </tr></table>
                        </td>
                    </tr></table>
                </td>
            </tr>
            <tr>
                <td class="cover-data-label">Número de Identificación<br/>del respondiente (ID):</td>
                <td><table><tr><td class="box id-box">{data.respondent_cedula}</td></tr></table></td>
            </tr>
        </table>

        <!-- Título verde centrado -->
        <div class="cover-main">
            <div class="green-title">
                CUESTIONARIO PARA LA EVALUACIÓN<br/>DEL ESTRÉS TERCERA VERSIÓN
            </div>
        </div>

        <!-- Logos pie de página -->
        <table class="cover-footer-table">
            <tr>
                <td class="footer-logo-left" style="text-align:left; padding-left:1cm;">
                    <img src="{img_colombia}" alt="Escudo de Colombia"><br/>
                    <span class="org-name">Ministerio de la Protección Social</span>
                    <span class="org-sub">República de Colombia</span>
                </td>
                <td class="footer-logo-right" style="text-align:right; padding-right:1cm;">
                    <img src="{img_javeriana}" alt="Logo Javeriana">
                </td>
            </tr>
        </table>
    </div>

    <!-- PÁGINA DE CONTENIDO -->
    <div class="content-page">

        <!-- Encabezado con logos -->
        <table class="header-table">
            <tr>
                <td class="logo-left-cell">
                    <table><tr>
                        <td><img src="{img_colombia}" alt="Escudo de Colombia"></td>
                        <td class="logo-text"><strong>Ministerio de la Protección Social</strong>República de Colombia</td>
                    </tr></table>
                </td>
                <td class="logo-right-cell">
                    <img src="{img_javeriana}" alt="Logo Javeriana">
                </td>
            </tr>
        </table>

        <div class="title">
            CUESTIONARIO PARA LA EVALUACIÓN DEL ESTRÉS – TERCERA VERSIÓN
        </div>

        <div class="instructions">
            <strong>Señale con una X la casilla que indique la frecuencia con que se le han presentado los siguientes malestares en los últimos tres meses.</strong>
        </div>

        <table class="questionnaire-table">
            <thead>
                <tr>
                    <th style="width:55%;">Malestares</th>
                    <th style="width:11.25%;">Siempre</th>
                    <th style="width:11.25%;">Casi<br/>siempre</th>
                    <th style="width:11.25%;">A veces</th>
                    <th style="width:11.25%;">Nunca</th>
                </tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>
    </div>

</body>
</html>"""


class PDFFromHTMLRequest(BaseModel):
    html: str
    filename: str


@app.post("/api/generate-pdf-server")
async def generate_pdf_server(data: PDFFromHTMLRequest):
    """Receive rendered HTML from the frontend and convert it to PDF using Playwright."""
    from playwright.async_api import async_playwright

    file_dir = os.path.join(RESULTS_PDF_DIR, "stress")
    if "extralaborales" in data.filename.lower():
        file_dir = os.path.join(RESULTS_PDF_DIR, "extralaborales")
    elif "intralaboral_a" in data.filename.lower() or "intralaborales-a" in data.filename.lower():
        file_dir = os.path.join(RESULTS_PDF_DIR, "intralaborales-a")
    elif "intralaboral_b" in data.filename.lower() or "intralaborales-b" in data.filename.lower():
        file_dir = os.path.join(RESULTS_PDF_DIR, "intralaborales-b")
        
    os.makedirs(file_dir, exist_ok=True)
    file_path = os.path.join(file_dir, data.filename)

    async with async_playwright() as p:
        browser = await p.chromium.launch()
        page = await browser.new_page()
        await page.set_content(data.html, wait_until="networkidle")
        await page.pdf(
            path=file_path,
            format="Letter",
            margin={"top": "1.01cm", "bottom": "1.01cm", "left": "1.01cm", "right": "1.01cm"},
            print_background=True,
        )
        await browser.close()

    return {"success": True, "filename": data.filename}


@app.get("/api/pdfs/stress")
async def list_generated_pdfs():
    """List all PDF files in the stress directory."""
    folder = os.path.join(RESULTS_PDF_DIR, "stress")
    if not os.path.exists(folder):
        return []
    
    files = []
    for f in os.listdir(folder):
        if f.lower().endswith(".pdf"):
            path = os.path.join(folder, f)
            size = os.path.getsize(path) / 1024  # KB
            created = os.path.getctime(path)
            files.append({
                "filename": f,
                "size_kb": round(size, 2),
                "created_at": datetime.fromtimestamp(created).strftime("%Y-%m-%d %H:%M:%S")
            })
    
    # Sort by newest first
    files.sort(key=lambda x: x["created_at"], reverse=True)
    return files


@app.get("/api/pdfs/stress/{filename}")
async def download_generated_pdf(filename: str):
    """Download a specific PDF file."""
    path = os.path.join(RESULTS_PDF_DIR, "stress", filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(path, media_type="application/pdf", filename=filename)


@app.delete("/api/pdfs/stress/{filename}")
async def delete_generated_pdf(filename: str):
    """Delete a specific PDF file."""
    path = os.path.join(RESULTS_PDF_DIR, "stress", filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    os.remove(path)
    return {"success": True, "message": f"Deleted {filename}"}


@app.post("/api/pdfs/stress/bulk-download")
async def bulk_download_pdfs(filenames: List[str]):
    """Download multiple PDF files as a ZIP archive."""
    import zipfile
    import io
    
    if not filenames:
        raise HTTPException(status_code=400, detail="No files specified")
    
    folder = os.path.join(RESULTS_PDF_DIR, "stress")
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for fname in filenames:
            file_path = os.path.join(folder, fname)
            if os.path.exists(file_path):
                zip_file.write(file_path, arcname=fname)
    
    zip_buffer.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        zip_buffer, 
        media_type="application/zip", 
        headers={"Content-Disposition": f"attachment; filename=reportes_estres_{timestamp}.zip"}
    )


@app.post("/api/pdfs/stress/bulk-delete")
async def bulk_delete_pdfs(filenames: List[str]):
    """Delete multiple PDF files."""
    deleted_count = 0
    errors = []
    
    folder = os.path.join(RESULTS_PDF_DIR, "stress")
    
    for fname in filenames:
        try:
            file_path = os.path.join(folder, fname)
            if os.path.exists(file_path):
                os.remove(file_path)
                deleted_count += 1
        except Exception as e:
            errors.append(f"{fname}: {str(e)}")
            
    return {
        "success": True, 
        "deleted_count": deleted_count, 
        "errors": errors
    }


# ==========================================
# EXTRALABORALES PDF ENDPOINTS
# ==========================================

# ==========================================
# EXTRALABORALES PDF ENDPOINTS
# ==========================================

@app.get("/api/pdfs/extralaborales")
async def list_generated_pdfs_extralaborales():
    """List all PDF files in the extralaborales directory."""
    folder = os.path.join(RESULTS_PDF_DIR, "extralaborales")
    if not os.path.exists(folder):
        return []
    
    files = []
    for f in os.listdir(folder):
        if f.lower().endswith(".pdf"):
            path = os.path.join(folder, f)
            size = os.path.getsize(path) / 1024  # KB
            created = os.path.getctime(path)
            files.append({
                "filename": f,
                "size_kb": round(size, 2),
                "created_at": datetime.fromtimestamp(created).strftime("%Y-%m-%d %H:%M:%S")
            })
    
    # Sort by newest first
    files.sort(key=lambda x: x["created_at"], reverse=True)
    return files


@app.get("/api/pdfs/extralaborales/{filename}")
async def download_generated_pdf_extralaborales(filename: str):
    """Download a specific PDF file."""
    path = os.path.join(RESULTS_PDF_DIR, "extralaborales", filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(path, media_type="application/pdf", filename=filename)


@app.delete("/api/pdfs/extralaborales/{filename}")
async def delete_generated_pdf_extralaborales(filename: str):
    """Delete a specific PDF file."""
    path = os.path.join(RESULTS_PDF_DIR, "extralaborales", filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    os.remove(path)
    return {"success": True, "message": f"Deleted {filename}"}


@app.post("/api/pdfs/extralaborales/bulk-download")
async def bulk_download_pdfs_extralaborales(filenames: List[str]):
    """Download multiple PDF files as a ZIP archive."""
    import zipfile
    import io
    
    if not filenames:
        raise HTTPException(status_code=400, detail="No files specified")
    
    folder = os.path.join(RESULTS_PDF_DIR, "extralaborales")
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for fname in filenames:
            file_path = os.path.join(folder, fname)
            if os.path.exists(file_path):
                zip_file.write(file_path, arcname=fname)
    
    zip_buffer.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        zip_buffer, 
        media_type="application/zip", 
        headers={"Content-Disposition": f"attachment; filename=reportes_extralaborales_{timestamp}.zip"}
    )


@app.post("/api/pdfs/extralaborales/bulk-delete")
async def bulk_delete_pdfs_extralaborales(filenames: List[str]):
    """Delete multiple PDF files."""
    deleted_count = 0
    errors = []
    
    folder = os.path.join(RESULTS_PDF_DIR, "extralaborales")
    
    for fname in filenames:
        try:
            file_path = os.path.join(folder, fname)
            if os.path.exists(file_path):
                os.remove(file_path)
                deleted_count += 1
        except Exception as e:
            errors.append(f"{fname}: {str(e)}")
            
    return {
        "success": True, 
        "deleted_count": deleted_count, 
        "errors": errors
    }


# ==========================================
# INTRALABORALES A PDF ENDPOINTS
# ==========================================

@app.get("/api/pdfs/intralaborales-a")
async def list_generated_pdfs_intralaborales_a():
    """List all PDF files in the intralaborales-a directory."""
    folder = os.path.join(RESULTS_PDF_DIR, "intralaborales-a")
    if not os.path.exists(folder):
        return []
    
    files = []
    for f in os.listdir(folder):
        if f.lower().endswith(".pdf"):
            path = os.path.join(folder, f)
            size = os.path.getsize(path) / 1024  # KB
            created = os.path.getctime(path)
            files.append({
                "filename": f,
                "size_kb": round(size, 2),
                "created_at": datetime.fromtimestamp(created).strftime("%Y-%m-%d %H:%M:%S")
            })
    
    # Sort by newest first
    files.sort(key=lambda x: x["created_at"], reverse=True)
    return files


@app.get("/api/pdfs/intralaborales-a/{filename}")
async def download_generated_pdf_intralaborales_a(filename: str):
    """Download a specific PDF file."""
    path = os.path.join(RESULTS_PDF_DIR, "intralaborales-a", filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(path, media_type="application/pdf", filename=filename)


@app.delete("/api/pdfs/intralaborales-a/{filename}")
async def delete_generated_pdf_intralaborales_a(filename: str):
    """Delete a specific PDF file."""
    path = os.path.join(RESULTS_PDF_DIR, "intralaborales-a", filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    os.remove(path)
    return {"success": True, "message": f"Deleted {filename}"}


@app.post("/api/pdfs/intralaborales-a/bulk-download")
async def bulk_download_pdfs_intralaborales_a(filenames: List[str]):
    """Download multiple PDF files as a ZIP archive."""
    import zipfile
    import io
    
    if not filenames:
        raise HTTPException(status_code=400, detail="No files specified")
    
    folder = os.path.join(RESULTS_PDF_DIR, "intralaborales-a")
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for fname in filenames:
            file_path = os.path.join(folder, fname)
            if os.path.exists(file_path):
                zip_file.write(file_path, arcname=fname)
    
    zip_buffer.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        zip_buffer, 
        media_type="application/zip", 
        headers={"Content-Disposition": f"attachment; filename=reportes_intralaborales_a_{timestamp}.zip"}
    )


@app.post("/api/pdfs/intralaborales-a/bulk-delete")
async def bulk_delete_pdfs_intralaborales_a(filenames: List[str]):
    """Delete multiple PDF files."""
    deleted_count = 0
    errors = []
    
    folder = os.path.join(RESULTS_PDF_DIR, "intralaborales-a")
    
    for fname in filenames:
        try:
            file_path = os.path.join(folder, fname)
            if os.path.exists(file_path):
                os.remove(file_path)
                deleted_count += 1
        except Exception as e:
            errors.append(f"{fname}: {str(e)}")
            
    return {
        "success": True, 
        "deleted_count": deleted_count, 
        "errors": errors
    }


@app.get("/api/pdfs/intralaborales-b")
async def list_generated_pdfs_intralaborales_b():
    """List all PDF files in the intralaborales-b directory."""
    folder = os.path.join(RESULTS_PDF_DIR, "intralaborales-b")
    if not os.path.exists(folder):
        return []
    
    files = []
    for f in os.listdir(folder):
        if f.lower().endswith(".pdf"):
            path = os.path.join(folder, f)
            size = os.path.getsize(path) / 1024  # KB
            created = os.path.getctime(path)
            files.append({
                "filename": f,
                "size_kb": round(size, 2),
                "created_at": datetime.fromtimestamp(created).strftime("%Y-%m-%d %H:%M:%S")
            })
    
    # Sort by newest first
    files.sort(key=lambda x: x["created_at"], reverse=True)
    return files


@app.get("/api/pdfs/intralaborales-b/{filename}")
async def download_generated_pdf_intralaborales_b(filename: str):
    """Download a specific PDF file."""
    path = os.path.join(RESULTS_PDF_DIR, "intralaborales-b", filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(path, media_type="application/pdf", filename=filename)


@app.delete("/api/pdfs/intralaborales-b/{filename}")
async def delete_generated_pdf_intralaborales_b(filename: str):
    """Delete a specific PDF file."""
    path = os.path.join(RESULTS_PDF_DIR, "intralaborales-b", filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    os.remove(path)
    return {"success": True, "message": f"Deleted {filename}"}


@app.post("/api/pdfs/intralaborales-b/bulk-download")
async def bulk_download_pdfs_intralaborales_b(filenames: List[str]):
    """Download multiple PDF files as a ZIP archive."""
    import zipfile
    import io
    
    if not filenames:
        raise HTTPException(status_code=400, detail="No files specified")
    
    folder = os.path.join(RESULTS_PDF_DIR, "intralaborales-b")
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for fname in filenames:
            file_path = os.path.join(folder, fname)
            if os.path.exists(file_path):
                zip_file.write(file_path, arcname=fname)
    
    zip_buffer.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        zip_buffer, 
        media_type="application/zip", 
        headers={"Content-Disposition": f"attachment; filename=reportes_intralaborales_b_{timestamp}.zip"}
    )


@app.post("/api/pdfs/intralaborales-b/bulk-delete")
async def bulk_delete_pdfs_intralaborales_b(filenames: List[str]):
    """Delete multiple PDF files."""
    deleted_count = 0
    errors = []
    
    folder = os.path.join(RESULTS_PDF_DIR, "intralaborales-b")
    
    for fname in filenames:
        try:
            file_path = os.path.join(folder, fname)
            if os.path.exists(file_path):
                os.remove(file_path)
                deleted_count += 1
        except Exception as e:
            errors.append(f"{fname}: {str(e)}")
            
    return {
        "success": True, 
        "deleted_count": deleted_count, 
        "errors": errors
    }


# CORS middleware


# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ==========================================
# CONSTANTS & CONFIGURATION
# ==========================================
# Sequence of questionnaires for the linear flow
SEQUENCE = ["datos-generales", "estres", "extralaborales", "intralaborales-a", "intralaborales-b"]

# ==========================================
# MODELS
# ==========================================

class Session(BaseModel):
    cedula: str
    name: Optional[str] = None
    completed_forms: List[str] = []
    current_step: str = "datos-generales"
    last_active: str

class QuestionResponse(BaseModel):
    question_id: int
    response_value: int


class SurveySubmission(BaseModel):
    questionnaire_id: str
    respondent_cedula: Optional[str] = None
    respondent_name: Optional[str] = None
    respondent_email: Optional[str] = None
    department: Optional[str] = None
    responses: List[QuestionResponse]

class FormSubmission(BaseModel):
    form_id: str
    submitted_at: Optional[str] = None
    data: Dict[str, Any]

# Helper functions
def ensure_dirs():
    """Ensure data and questionnaires directories exist"""
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(QUESTIONNAIRES_DIR, exist_ok=True)

def get_sessions_file() -> str:
    """Get the sessions file path"""
    ensure_dirs()
    return os.path.join(DATA_DIR, "sessions.json")

def load_sessions() -> Dict[str, dict]:
    """Load all sessions"""
    file_path = get_sessions_file()
    if not os.path.exists(file_path):
        return {}
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, FileNotFoundError):
        return {}

def save_session(session_data: dict):
    """Save/Update a session"""
    file_path = get_sessions_file()
    sessions = load_sessions()
    
    # Update specific session
    sessions[session_data["cedula"]] = session_data
    
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(sessions, f, ensure_ascii=False, indent=2)

def get_next_step(completed_forms: List[str]) -> str:
    """Determine the next step based on completed starts"""
    for step in SEQUENCE:
        if step not in completed_forms:
            return step
    return "completed"

def load_questionnaire(questionnaire_id: str) -> Dict[str, Any]:

    """Load a questionnaire by ID"""
    file_path = os.path.join(QUESTIONNAIRES_DIR, f"{questionnaire_id}.json")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"Questionnaire '{questionnaire_id}' not found")
    
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def get_all_questionnaires() -> List[Dict[str, Any]]:
    """Get list of all available questionnaires"""
    ensure_dirs()
    questionnaires = []
    
    for file_path in glob.glob(os.path.join(QUESTIONNAIRES_DIR, "*.json")):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                questionnaires.append({
                    "id": data.get("id"),
                    "name": data.get("name"),
                    "description": data.get("description"),
                    "version": data.get("version"),
                    "icon": data.get("icon", "📋"),
                    "color": data.get("color", "#6366f1"),
                    "question_count": len(data.get("questions", []))
                })
        except (json.JSONDecodeError, KeyError) as e:
            print(f"Error loading questionnaire {file_path}: {e}")
    
    return questionnaires

def get_responses_file(questionnaire_id: str) -> str:
    """Get the responses file path for a questionnaire"""
    ensure_dirs()
    return os.path.join(DATA_DIR, f"responses_{questionnaire_id}.json")

def load_responses(questionnaire_id: str) -> List[dict]:
    """Load all saved responses for a questionnaire"""
    file_path = get_responses_file(questionnaire_id)
    if not os.path.exists(file_path):
        return []
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, FileNotFoundError):
        return []

def save_response(questionnaire_id: str, response_data: dict):
    """Save a new response for a questionnaire"""
    file_path = get_responses_file(questionnaire_id)
    responses = load_responses(questionnaire_id)
    responses.append(response_data)
    
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(responses, f, ensure_ascii=False, indent=2)

def get_form_responses_file(form_id: str) -> str:
    """Get the responses file path for a form"""
    ensure_dirs()
    return os.path.join(DATA_DIR, f"form_{form_id}.json")

def load_form_responses(form_id: str) -> List[dict]:
    """Load all saved responses for a form"""
    file_path = get_form_responses_file(form_id)
    if not os.path.exists(file_path):
        return []
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, FileNotFoundError):
        return []

def save_form_response(form_id: str, response_data: dict):
    """Save a new response for a form"""
    file_path = get_form_responses_file(form_id)
    responses = load_form_responses(form_id)
    responses.append(response_data)
    
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(responses, f, ensure_ascii=False, indent=2)

# API Endpoints
@app.get("/api/questionnaires")
async def list_questionnaires():
    """Get all available questionnaires"""
    questionnaires = get_all_questionnaires()
    return {
        "questionnaires": questionnaires,
        "total": len(questionnaires)
    }

@app.get("/api/questionnaires/{questionnaire_id}")
async def get_questionnaire(questionnaire_id: str):
    """Get a specific questionnaire with its questions and options"""
    data = load_questionnaire(questionnaire_id)
    return {
        "id": data.get("id"),
        "name": data.get("name"),
        "short_name": data.get("short_name"),
        "description": data.get("description"),
        "version": data.get("version"),
        "icon": data.get("icon", "📋"),
        "color": data.get("color", "#6366f1"),
        "estimated_time": data.get("estimated_time"),
        "options": data.get("options", []),
        "questions": data.get("questions", []),
        "sections": data.get("sections", []),
        "conditional_questions": data.get("conditional_questions", []),
        "total_questions": len(data.get("questions", []))
    }

@app.post("/api/submit")
async def submit_survey(submission: SurveySubmission):
    """Submit survey responses"""
    # Load questionnaire to validate
    questionnaire = load_questionnaire(submission.questionnaire_id)
    
    valid_question_ids = {q["id"] for q in questionnaire.get("questions", [])}
    valid_response_values = {opt["value"] for opt in questionnaire.get("options", [])}
    
    # Get non-conditional question IDs (required questions)
    required_question_ids = {q["id"] for q in questionnaire.get("questions", []) if not q.get("conditional", False)}
    
    # Validate responses
    for response in submission.responses:
        if response.question_id not in valid_question_ids:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid question ID: {response.question_id}"
            )
        if response.response_value not in valid_response_values:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid response value: {response.response_value}"
            )
    
    # Check all required (non-conditional) questions are answered
    answered_ids = {r.question_id for r in submission.responses}
    missing_required = required_question_ids - answered_ids
    if missing_required:
        raise HTTPException(
            status_code=400,
            detail=f"Missing responses for required questions: {sorted(missing_required)}"
        )
    
    # Create response record
    response_id = str(uuid.uuid4())
    questions_dict = {q["id"]: q for q in questionnaire.get("questions", [])}
    options_dict = {opt["value"]: opt for opt in questionnaire.get("options", [])}
    
    response_data = {
        "id": response_id,
        "submitted_at": datetime.now().isoformat(),
        "respondent_cedula": submission.respondent_cedula,
        "respondent_name": submission.respondent_name,
        "respondent_email": submission.respondent_email,
        "department": submission.department,
        "responses": [
            {
                "question_id": r.question_id,
                "question_text": questions_dict[r.question_id]["text"],
                "response_value": r.response_value,
                "response_label": options_dict[r.response_value]["label"]
            }
            for r in submission.responses
        ]
    }
    
    # Save to file
    save_response(submission.questionnaire_id, response_data)
    
    # --- SESSION UPDATE LOGIC ---
    if submission.respondent_cedula:
        sessions = load_sessions()
        existing_session = sessions.get(submission.respondent_cedula)
        
        if existing_session:
            completed = set(existing_session.get("completed_forms", []))
            completed.add(submission.questionnaire_id)
            
            existing_session["completed_forms"] = list(completed)
            existing_session["last_active"] = datetime.now().isoformat()
            
            # Determine next step
            next_step = get_next_step(list(completed))
            existing_session["current_step"] = next_step
            
            save_session(existing_session)
            
            return {
                "success": True,
                "message": "Encuesta enviada exitosamente",
                "submission_id": response_id,
                "next_step": next_step
            }
    
    return {
        "success": True,
        "message": "Encuesta enviada exitosamente",
        "submission_id": response_id
    }

@app.get("/api/responses/{questionnaire_id}")
async def get_all_responses(questionnaire_id: str):
    """Get all saved responses for a questionnaire"""
    # Verify questionnaire exists
    load_questionnaire(questionnaire_id)
    
    responses = load_responses(questionnaire_id)
    return {
        "questionnaire_id": questionnaire_id,
        "responses": responses,
        "total": len(responses)
    }

@app.post("/api/submit-form")
async def submit_form(submission: FormSubmission):
    """Submit form data (datos generales)"""
    response_id = str(uuid.uuid4())
    
    response_data = {
        "id": response_id,
        "submitted_at": submission.submitted_at or datetime.now().isoformat(),
        "data": submission.data
    }
    
    # Save to file
    save_form_response(submission.form_id, response_data)
    
    # --- SESSION UPDATE LOGIC ---
    if submission.form_id == "datos-generales":
        cedula = submission.data.get("numero_identificacion")
        nombre = submission.data.get("nombre_completo")
        
        if cedula:
            sessions = load_sessions()
            existing_session = sessions.get(cedula, {})
            
            # Merge completed forms
            completed = set(existing_session.get("completed_forms", []))
            completed.add(submission.form_id)
            
            # Logic for Intralaboral Form Selection
            # If has personnel (Si) -> Do Intralaboral A (Skip B)
            # If no personnel (No) -> Do Intralaboral B (Skip A)
            personal_cargo = submission.data.get("tiene_personal_cargo")
            
            if personal_cargo == "si":
                completed.add("intralaborales-b")
                if "intralaborales-a" in completed:
                    completed.remove("intralaborales-a")
            elif personal_cargo == "no":
                completed.add("intralaborales-a")
                if "intralaborales-b" in completed:
                    completed.remove("intralaborales-b")
            
            new_session = {
                "cedula": cedula,
                "name": nombre or existing_session.get("name"),
                "completed_forms": list(completed),
                "last_active": datetime.now().isoformat(),
            }
            # Determine next step (re-calculation)
            new_session["current_step"] = get_next_step(new_session["completed_forms"])
            
            save_session(new_session)
            
            return {
                "success": True,
                "message": "Datos guardados exitosamente",
                "submission_id": response_id,
                "next_step": new_session["current_step"]
            }

    return {
        "success": True,
        "message": "Datos guardados exitosamente",
        "submission_id": response_id
    }

@app.get("/api/form-responses/{form_id}")
async def get_form_responses(form_id: str):
    """Get all saved responses for a form"""
    responses = load_form_responses(form_id)
    return {
        "form_id": form_id,
        "responses": responses,
        "total": len(responses)
    }

@app.get("/api/lookup-cedula/{cedula}")
async def lookup_cedula(cedula: str):
    """Look up respondent data by cedula from datos-generales form"""
    responses = load_form_responses("datos-generales")
    
    for response in responses:
        data = response.get("data", {})
        if data.get("numero_identificacion") == cedula:
            return {
                "found": True,
                "cedula": cedula,
                "nombre": data.get("nombre_completo"),
                "departamento": data.get("departamento_area"),
                "cargo": data.get("nombre_cargo"),
                "tipo_cargo": data.get("tipo_cargo")
            }
    
    return {
        "found": False,
        "cedula": cedula,
        "nombre": None,
        "departamento": None,
        "cargo": None,
        "tipo_cargo": None
    }

@app.get("/api/analysis-report")
async def get_analysis_report():
    """Get the full analysis report for all questionnaires"""
    engine = AnalysisEngine(DATA_DIR, QUESTIONNAIRES_DIR)
    return engine.get_global_report()

@app.post("/api/ai-analysis")
async def generate_ai_analysis(data: Dict[str, Any]):
    """
    Sends data to n8n webhook for LLM processing.
    """
    webhook_url = os.getenv("N8N_WEBHOOK_URL")
    
    if webhook_url:
        import httpx
        try:
            async with httpx.AsyncClient() as client:
                response = await client.post(webhook_url, json=data, timeout=60.0)
                if response.status_code == 200:
                    return response.json()
        except Exception as e:
            print(f"Error calling n8n webhook: {e}")
            
    # Fallback to simulation if no webhook or if it fails
    intra = data.get("questionnaires", {}).get("intralaborales-a", {})
    return {
        "summary": f"El análisis revela un nivel de riesgo {intra.get('risk_level')} en el factor intralaboral. Se destaca que el dominio de Liderazgo muestra puntuaciones favorables.",
        "recommendations": [
            "Implementar talleres de liderazgo participativo",
            "Realizar jornadas de bienestar enfocadas en manejo del estrés",
            "Optimizar la distribución de cargas laborales en departamentos críticos"
        ]
    }

@app.get("/api/statistics/{questionnaire_id}")
async def get_statistics(questionnaire_id: str):
    """Get survey statistics for a questionnaire"""
    questionnaire = load_questionnaire(questionnaire_id)
    responses = load_responses(questionnaire_id)
    
    if not responses:
        return {
            "questionnaire_id": questionnaire_id,
            "questionnaire_name": questionnaire.get("name"),
            "total_responses": 0,
            "question_stats": {},
            "options": questionnaire.get("options", [])
        }
    
    # Calculate statistics per question
    question_stats = {}
    options = questionnaire.get("options", [])
    option_values = [opt["value"] for opt in options]
    
    for question in questionnaire.get("questions", []):
        q_id = question["id"]
        values = []
        for response in responses:
            for r in response["responses"]:
                if r["question_id"] == q_id:
                    values.append(r["response_value"])
        
        if values:
            avg = sum(values) / len(values)
            distribution = {v: values.count(v) for v in option_values}
            
            question_stats[q_id] = {
                "question_text": question["text"],
                "category": question.get("category", ""),
                "average": round(avg, 2),
                "total_responses": len(values),
                "distribution": distribution
            }
    
    return {
        "questionnaire_id": questionnaire_id,
        "questionnaire_name": questionnaire.get("name"),
        "total_responses": len(responses),
        "question_stats": question_stats,
        "options": questionnaire.get("options", [])
    }

@app.get("/api/export/excel/{questionnaire_id}")
async def export_excel(questionnaire_id: str):
    """Export questionnaire responses to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    # Load questionnaire and responses
    questionnaire = load_questionnaire(questionnaire_id)
    responses = load_responses(questionnaire_id)
    
    if not responses:
        raise HTTPException(status_code=404, detail="No responses found")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Respuestas"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Build headers
    questions = questionnaire.get("questions", [])
    headers = ["ID", "Fecha", "Cédula", "Nombre", "Departamento"]
    for q in questions:
        headers.append(f"P{q['id']}: {q['text'][:50]}...")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, response in enumerate(responses, 2):
        # Build response dict for quick lookup
        response_dict = {r["question_id"]: r["response_value"] for r in response.get("responses", [])}
        
        # Basic info
        ws.cell(row=row_idx, column=1, value=response.get("id", "")[:8])
        ws.cell(row=row_idx, column=2, value=response.get("submitted_at", "")[:10])
        ws.cell(row=row_idx, column=3, value=response.get("respondent_cedula", ""))
        ws.cell(row=row_idx, column=4, value=response.get("respondent_name", ""))
        ws.cell(row=row_idx, column=5, value=response.get("department", ""))
        
        # Question responses
        for col_idx, q in enumerate(questions, 6):
            value = response_dict.get(q["id"], "")
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"respuestas_{questionnaire_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/auth-config")
async def get_auth_config():
    """Get authentication configuration"""
    return {
        "allowed_user_id": os.getenv("ALLOWED_USER_ID"),
        "allowed_cedula": os.getenv("ALLOWED_CEDULA"),
        "allowed_email": os.getenv("ALLOWED_EMAIL")
    }

# --- SESSION ENDPOINTS ---

@app.get("/api/session/{cedula}")
async def get_session(cedula: str):
    """Get session status for a user"""
    sessions = load_sessions()
    session = sessions.get(cedula)
    
    if session:
        # Calculate progress
        completed_count = len(session.get("completed_forms", []))
        total_forms = len(SEQUENCE)
        progress = int((completed_count / total_forms) * 100)
        
        return {
            "found": True,
            "cedula": session["cedula"],
            "name": session.get("name"),
            "completed_forms": session.get("completed_forms", []),
            "current_step": session.get("current_step"),
            "progress": progress,
            "finished": session.get("current_step") == "completed"
        }
    
    return {"found": False}


# Serve frontend static files
@app.get("/")
async def serve_index():
    """Serve the questionnaire selector page"""
    return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))

@app.get("/encuesta/{questionnaire_id}")
async def serve_survey(questionnaire_id: str):
    """Serve the survey page for a specific questionnaire"""
    return FileResponse(os.path.join(FRONTEND_DIR, "encuesta.html"))

@app.get("/resultados/{questionnaire_id}")
async def serve_results(questionnaire_id: str):
    """Serve the results page for a specific questionnaire"""
    return FileResponse(os.path.join(FRONTEND_DIR, "resultados.html"))

@app.get("/ficha-datos")
async def serve_ficha_datos():
    """Serve the ficha de datos generales page"""
    return FileResponse(os.path.join(FRONTEND_DIR, "ficha-datos.html"))

@app.get("/resultados-dashboard")
async def serve_results_dashboard():
    """Serve the results dashboard page"""
    return FileResponse(os.path.join(FRONTEND_DIR, "resultados-dashboard.html"))

@app.get("/reporte-global")
async def serve_global_report():
    """Serve the global analysis report page"""
    return FileResponse(os.path.join(FRONTEND_DIR, "reporte-global.html"))

@app.get("/reporte-premium")
async def serve_premium_report():
    """Serve the premium multi-page report"""
    return FileResponse(os.path.join(FRONTEND_DIR, "reporte-premium.html"))

# Legacy routes for backwards compatibility
@app.get("/encuesta")
async def serve_survey_legacy():
    """Redirect to selector if no questionnaire specified"""
    return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))

@app.get("/resultados")
async def serve_results_legacy():
    """Redirect to selector if no questionnaire specified"""
    return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))

# Mount static files (CSS, JS) - must be after API routes
app.mount("/", StaticFiles(directory=FRONTEND_DIR, html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    import argparse

    parser = argparse.ArgumentParser(description="Iniciar servidor de Formularios")
    parser.add_argument("--port", type=int, default=8000, help="Puerto del servidor (default: 8000)")
    parser.add_argument("--host", type=str, default="0.0.0.0", help="Host del servidor (default: 0.0.0.0)")
    args = parser.parse_args()

    ensure_dirs()
    print("\n" + "="*60)
    print(f"Sistema de Cuestionarios iniciado en puerto {args.port}!")
    print("="*60)
    print("\n SELECTOR DE CUESTIONARIOS:")
    print(f"   http://localhost:{args.port}/")
    print("\n ENCUESTA (ejemplo):")
    print(f"   http://localhost:{args.port}/encuesta/estres")
    print("\n RESULTADOS (ejemplo):")
    print(f"   http://localhost:{args.port}/resultados/estres")
    print(" IMPORTANTE: No cerrar esta ventana mientras se usen los cuestionarios")
    print("\n Agrega nuevos cuestionarios en:")
    print(f"   {QUESTIONNAIRES_DIR}")
    print("="*60 + "\n")
    
    uvicorn.run(app, host=args.host, port=args.port)

    print("="*60 + "\n")


